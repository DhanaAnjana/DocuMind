"""Document processing utilities."""
from typing import List, Union
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
import tiktoken
from langchain.text_splitter import RecursiveCharacterTextSplitter

def extract_text(file_path: str, content_type: str) -> str:
    """Extract text content from various document types."""
    if content_type == "application/pdf":
        return _extract_from_pdf(file_path)
    elif content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return _extract_from_docx(file_path)
    elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _extract_from_xlsx(file_path)
    elif content_type.startswith("text/"):
        return _extract_from_text(file_path)
    else:
        raise ValueError(f"Unsupported content type: {content_type}")

def _extract_from_pdf(file_path: str) -> str:
    """Extract text from PDF file."""
    reader = PdfReader(file_path)
    return " ".join(page.extract_text() for page in reader.pages)

def _extract_from_docx(file_path: str) -> str:
    """Extract text from DOCX file."""
    doc = Document(file_path)
    return " ".join(paragraph.text for paragraph in doc.paragraphs)

def _extract_from_xlsx(file_path: str) -> str:
    """Extract text from XLSX file."""
    wb = openpyxl.load_workbook(file_path)
    texts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            texts.extend(str(cell.value) for cell in row if cell.value)
    return " ".join(texts)

def _extract_from_text(file_path: str) -> str:
    """Extract text from plain text file."""
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def split_text(text: str) -> List[str]:
    """Split text into chunks using LangChain's text splitter."""
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=1000,
        chunk_overlap=200,
        length_function=lambda x: len(tiktoken.get_encoding("cl100k_base").encode(x)),
        separators=["\n\n", "\n", ". ", " ", ""]
    )
    return text_splitter.split_text(text)